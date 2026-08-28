import pandas as pd

from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment

def _style(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

def create_excel(results):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        results["por_crianca"].to_excel(
            writer, sheet_name="Base_por_Crianca", index=False
        )
        results["resumo_categoria"].to_excel(
            writer, sheet_name="Categoria_Amostragem"
        )
        results["resumo_participacao"].to_excel(
            writer, sheet_name="Participacao_Pote"
        )
        results["resumo_participacao_lamina"].to_excel(
            writer, sheet_name="Participacao_Lamina"
        )
        results["resumo_especies"].to_excel(
            writer, sheet_name="Prevalencia_por_Especie"
        )
        results["resumo_poli"].to_excel(
            writer, sheet_name="Poliparasitismo", index=False
        )
        results["resumo_metodos"].to_excel(
            writer, sheet_name="Comparacao_Metodos", index=False
        )
        results["tabela_cruzada"].to_excel(
            writer, sheet_name="Especie_x_Metodo"
        )
        results["prev_por_n_coletas"].to_excel(
            writer, sheet_name="Prevalencia_x_NColetas"
        )

        wb = writer.book
        for ws in wb.worksheets:
            _style(ws)

    output.seek(0)
    return output.getvalue()
